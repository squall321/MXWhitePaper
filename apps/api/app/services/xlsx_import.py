"""xlsx (Excel) → DocumentJSON v1.0.

docx/pptx import 와 같은 계약: ``{document, summary}`` 반환. 시트 1개를
섹션 1개로 만들고, 시트의 표를 (크기 무관) TableBlock 으로,
embedded 차트는 ChartBlock 으로 분배한다. 본문 walk 가 끝나면 docx/pptx 와
동일하게 ``apply_widget_markers`` + ``apply_widget_autodetect`` 를 태운다.
표는 기본적으로 TableBlock 으로 보존되며, autodetect 가 모양을 보고
label/value 헤더 → KpiCardsBlock, name/start/end 헤더 → GanttBlock 으로만
자동 승격한다 (일반 숫자 표는 표 그대로 — 차트는 embedded 엑셀 차트나
``Widget: chart`` 마커로 생성). embedded 차트는 아래에서 직접 ChartBlock 으로.

openpyxl 은 ``data_only=True`` 로 열어 수식 대신 캐시된 계산값을 읽는다 —
수식 문자열을 그대로 넣으면 표가 숫자 표로 인식되지 않아 autodetect 가
동작하지 않기 때문.
"""
from __future__ import annotations

import io
import zipfile
from typing import Any

import ulid

from .docx_import import ImportSummary

# 셀 텍스트 캡 — 비정상적으로 긴 셀로 문서가 비대해지는 것 방지.
_CELL_MAX = 2000


def _new_id() -> str:
    return str(ulid.new())


def is_xlsx_zip_magic(buf: bytes) -> bool:
    """PK zip 매직만 확인 (docx/pptx 와 동일 — xlsx 도 OOXML zip)."""
    return len(buf) >= 2 and buf[:2] == b"PK"


def is_xlsx_content(buf: bytes) -> bool:
    """zip 안에 xl/workbook.xml 이 있는지까지 확인 (매직만으로는 부족)."""
    try:
        with zipfile.ZipFile(io.BytesIO(buf)) as zf:
            names = set(zf.namelist())
        return "xl/workbook.xml" in names
    except (zipfile.BadZipFile, OSError):
        return False


def _settings_default_division() -> str:
    from app.core.config import get_settings

    return get_settings().import_default_division


def _settings_default_confidentiality() -> str:
    from app.core.config import get_settings

    return get_settings().import_default_confidentiality


def _cell_text(value: Any) -> str:
    """openpyxl 셀 값 → 표시 문자열. None 은 빈 문자열, 나머지는 str()."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        # 12.0 → "12" (정수로 떨어지는 float 은 소수점 제거 — 표가 깔끔)
        return str(int(value))
    return str(value)[:_CELL_MAX]


def _used_rows(ws: Any) -> list[list[str]]:
    """시트의 used range 를 2D 문자열 리스트로. 완전히 빈 trailing 행/열은
    제거한다 (openpyxl 의 max_row/max_col 이 과대평가하는 경우 방지)."""
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_text(v) for v in row])
    # trailing 빈 행 제거
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        return []
    # trailing 빈 열 제거 (모든 행에서 비어있는 우측 열)
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    while width > 0 and all(r[width - 1] == "" for r in rows):
        for r in rows:
            r.pop()
        width -= 1
    return rows


def _table_block_from_rows(rows: list[list[str]]) -> dict[str, Any]:
    """첫 행을 헤더로 보는 TableBlock. label/value · name/start/end 형태면
    후속 autodetect 가 kpi/gantt 로 승격할 수 있게 docx 의 fast-path 와
    동일한 shape 사용 (일반 표는 그대로 유지)."""
    return {
        "type": "table",
        "id": _new_id(),
        "headers": rows[0],
        "rows": rows[1:],
    }


# openpyxl chart 타입 → ChartBlock.chartType. 미지원은 bar fallback.
_CHART_TYPE_MAP = {
    "BarChart": "bar",
    "LineChart": "line",
    "PieChart": "pie",
    "ScatterChart": "scatter",
    "AreaChart": "area",
}


def _chart_block_from_ws_chart(chart: Any, summary: ImportSummary) -> dict[str, Any] | None:
    """openpyxl 차트 객체 → ChartBlock. 추출 실패 시 None + warning."""
    try:
        chart_type = _CHART_TYPE_MAP.get(type(chart).__name__, "bar")
        labels: list[str] = []
        series_out: list[dict[str, Any]] = []
        for s_idx, ser in enumerate(getattr(chart, "series", []) or []):
            values: list[float] = []
            numref = getattr(ser, "val", None)
            pts = getattr(getattr(numref, "numRef", None), "numCache", None)
            if pts is not None:
                for pt in getattr(pts, "pt", []) or []:
                    try:
                        values.append(float(pt.v))
                    except (TypeError, ValueError):
                        values.append(0.0)
            name = None
            tx = getattr(ser, "tx", None)
            if tx is not None and getattr(tx, "strRef", None) is not None:
                cache = getattr(tx.strRef, "strCache", None)
                if cache is not None and getattr(cache, "pt", None):
                    name = cache.pt[0].v
            series_out.append({"name": name or f"Series {s_idx + 1}", "data": values})
        # 카테고리 라벨 (첫 시리즈의 cat 캐시)
        first = (getattr(chart, "series", []) or [None])[0]
        if first is not None:
            cat = getattr(first, "cat", None)
            ref = getattr(cat, "strRef", None) or getattr(cat, "numRef", None)
            cache = getattr(ref, "strCache", None) or getattr(ref, "numCache", None)
            if cache is not None:
                labels = [str(pt.v) for pt in (getattr(cache, "pt", []) or [])]
        if not series_out:
            return None
        return {
            "type": "chart",
            "id": _new_id(),
            "chartType": chart_type,
            "data": {"labels": labels, "series": series_out},
        }
    except Exception as e:  # noqa: BLE001 — best-effort; never break import
        summary.warnings.append(f"차트 추출 실패 (원본 표는 유지): {e}")
        return None


def xlsx_to_document(
    buf: bytes,
    *,
    slug: str,
    title: str = "",
    owner_user_id: str | None = None,
) -> dict[str, Any]:
    """Top-level entry point — .xlsx 바이트를 DocumentJSON 으로."""
    if not is_xlsx_zip_magic(buf):
        raise ValueError("not a valid zip (.xlsx must be PK zip)")
    if not is_xlsx_content(buf):
        raise ValueError("zip does not contain xl/workbook.xml")

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(buf), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"xlsx 를 열 수 없습니다: {e}") from e

    summary = ImportSummary()
    sections: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        rows = _used_rows(ws)
        charts = list(getattr(ws, "_charts", []) or [])
        if not rows and not charts:
            summary.warnings.append(f"빈 시트 건너뜀: {ws.title}")
            continue

        blocks: list[dict[str, Any]] = []
        if rows:
            # 표는 크기와 무관하게 TableBlock(무제한 정적 표)으로 보존한다.
            # SpreadsheetBlock 은 에디터 cap(200행/26열)이 있어 큰 시트는 담을 수
            # 없고, import 가 SpreadsheetBlock 을 만들지 않는다는 컨벤션과도 일치
            # (lat storage gotcha 9). label/value·name/start/end 모양이면 아래
            # autodetect 가 kpi/gantt 로 승격한다.
            blocks.append(_table_block_from_rows(rows))
            summary.tables += 1
        for chart in charts:
            cb = _chart_block_from_ws_chart(chart, summary)
            if cb is not None:
                blocks.append(cb)

        sections.append(
            {
                "id": _new_id(),
                "level": 1,
                "title": ws.title,
                "blocks": blocks,
            }
        )

    if not sections:
        # 모든 시트가 비었어도 유효한 빈 문서는 반환 (사용자가 채울 수 있게)
        sections.append(
            {"id": _new_id(), "level": 1, "title": "Sheet1", "blocks": []}
        )

    # Widget post-pass — docx/pptx 와 동일. label/value→kpi, name/start/end→gantt
    # 모양만 자동 승격 (일반 표는 그대로).
    from . import widget_markers as _wm

    _wm.apply_widget_markers(sections, summary)
    _wm.apply_widget_autodetect(sections, summary)

    metadata: dict[str, Any] = {
        "division": _settings_default_division(),
        "owners": [owner_user_id] if owner_user_id else [],
        "tags": [],
        "confidentiality": _settings_default_confidentiality(),
    }
    doc: dict[str, Any] = {
        "schema_version": "1.0",
        "id": str(ulid.new()),
        "slug": slug,
        "title": (title.strip() or wb.worksheets[0].title if wb.worksheets else "Untitled")[:200],
        "metadata": metadata,
        "infobox": {},
        "sections": sections,
    }
    return {"document": doc, "summary": summary}
